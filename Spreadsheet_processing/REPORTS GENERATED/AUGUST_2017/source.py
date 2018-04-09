#Metrics Report -Version 1.0

from Tkinter import *
import tkFileDialog
import os
import openpyxl
path=path1='a'
def fun():
    global path
    path=tkFileDialog.askopenfilename()
    print("loaded export file")

def fun1():
    global path1
    path1=tkFileDialog.askopenfilename()
    print("loaded template file")

def main():
    print('Processing.......')
    global path
    global path1
    #os.chdir('C:')
    inbook=openpyxl.load_workbook(path)
    outbook=openpyxl.load_workbook(path1)
    insheet=inbook.get_sheet_by_name('Sheet1')
    
    outsheet=outbook.get_sheet_by_name('Tickets Data')
    
#print(insheet.max_column)   11
#print(insheet.max_row)         4
    x=11

    for i in range(2,insheet.max_row+1):
        outsheet.cell(row=i+x,column=1).value=insheet.cell(row=i,column=1).value                         #IM number
        outsheet.cell(row=i+x,column=2).value="Production Support"                                       #Problem Ticket Type
        outsheet.cell(row=i+x,column=3).value="Severity - "+str(insheet.cell(row=i,column=5).value)      #Ticket Priority
        outsheet.cell(row=i+x,column=4).value=insheet.cell(row=i,column=2) .value                        #Open Time
        outsheet.cell(row=i+x,column=8).value=insheet.cell(row=i,column=11) .value                       #Resolved Time
        outsheet.cell(row=i+x,column=9).value=insheet.cell(row=i,column=12) .value                       #Closed Time
        outsheet.cell(row=i+x,column=10).value=insheet.cell(row=i,column=3) .value                       #Status
        outsheet.cell(row=i+x,column=11).value='N'
        outsheet.cell(row=i+x,column=12).value='Y'
        outsheet.cell(row=i+x,column=13).value='Y'
        outsheet.cell(row=i+x,column=15).value=insheet.cell(row=i,column=4) .value                       #Problem Desc
        outsheet.cell(row=i+x,column=18).value='Non TAC'
        outsheet.cell(row=i+x,column=19).value='NA'
        outsheet.cell(row=i+x,column=20).value='NA'
    
        
    outbook.save('pleasebegood.xlsx')
    inbook.close()
    print('Processed  successfully !...')
    

root=Tk()
root.title('Metrics Report')
m=Button(root,text='EXPORT',command=fun)
m1=Button(root,text='TEMPLATE',command=fun1)
m2=Button(root,text='PROCESS',command=main)
m.pack(padx=50,pady=50)
m1.pack(padx=50,pady=50)
m2.pack(padx=50,pady=50)
root.mainloop()
